"""Exportación de reportes a Excel con openpyxl."""

from pathlib import Path
from typing import Any, Dict, Optional

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from config import ETIQUETAS_METODO_PAGO, MARCA_COLORES, RESTAURANTE, RUTA_EXPORTACION, RUTA_LOGO_PNG
from reports.utilidades_reporte import (
    ENCABEZADOS_DETALLE_DIARIO,
    detalle_con_separadores_factura,
    etiqueta_metodo_pago,
    numero_factura_corto,
    texto_comprador,
)

_NOMBRE_HOJA = "Ventas"
_FILA_ENCABEZADO = 1
_FILL_ENCABEZADO = PatternFill("solid", fgColor=MARCA_COLORES["fondo_tabla"].lstrip("#"))
_FONT_ENCABEZADO = Font(bold=True, color=MARCA_COLORES["naranja_oscuro"].lstrip("#"))
_FONT_TITULO = Font(bold=True, size=14, color=MARCA_COLORES["naranja_oscuro"].lstrip("#"))
_FONT_SUBTITULO = Font(color=MARCA_COLORES["texto_suave"].lstrip("#"))
_TAMANO_LOGO_PX = 56
_FILAS_RESERVADAS_LOGO = 4


def _asegurar_directorio_destino(ruta: Path) -> None:
    """Crea el directorio padre del archivo de destino si no existe."""
    ruta.parent.mkdir(parents=True, exist_ok=True)


def _ajustar_ancho_columnas(hoja) -> None:
    """Ajusta el ancho de columnas según el contenido."""
    for columna in hoja.columns:
        maximo = 0
        letra = get_column_letter(columna[0].column)
        for celda in columna:
            if celda.value is not None:
                maximo = max(maximo, len(str(celda.value)))
        hoja.column_dimensions[letra].width = min(max(maximo + 2, 12), 40)


def _insertar_logo(hoja) -> int:
    """
    Inserta el logo Hogareños en la hoja si existe el PNG.
    Retorna la fila donde comienza el texto del encabezado.
    """
    if not RUTA_LOGO_PNG.is_file():
        return 1

    imagen = XLImage(str(RUTA_LOGO_PNG))
    imagen.width = _TAMANO_LOGO_PX
    imagen.height = _TAMANO_LOGO_PX
    hoja.add_image(imagen, "A1")
    hoja.row_dimensions[1].height = 44
    return _FILAS_RESERVADAS_LOGO


def _escribir_encabezado_hoja(hoja, titulo: str, subtitulo: str) -> int:
    """Escribe logo, metadatos del reporte y retorna la fila donde empiezan los datos."""
    fila = _insertar_logo(hoja)

    hoja.cell(row=fila, column=1, value=RESTAURANTE["nombre"])
    hoja.cell(row=fila, column=1).font = _FONT_TITULO
    fila += 1

    hoja.cell(row=fila, column=1, value=RESTAURANTE["direccion"])
    hoja.cell(row=fila, column=1).font = _FONT_SUBTITULO
    fila += 1

    hoja.cell(row=fila, column=1, value=titulo)
    hoja.cell(row=fila, column=1).font = Font(
        bold=True, color=MARCA_COLORES["texto"].lstrip("#")
    )
    fila += 1

    hoja.cell(row=fila, column=1, value=subtitulo)
    hoja.cell(row=fila, column=1).font = _FONT_SUBTITULO
    return fila + 2


def _aplicar_estilo_encabezado_tabla(hoja, fila: int, columnas: int) -> None:
    """Aplica estilo a la fila de encabezados de la tabla de datos."""
    for col in range(1, columnas + 1):
        celda = hoja.cell(row=fila, column=col)
        celda.font = _FONT_ENCABEZADO
        celda.fill = _FILL_ENCABEZADO
        celda.alignment = Alignment(horizontal="center")


def _ajustar_columnas_detalle_diario(hoja) -> None:
    """Fija anchos legibles para el detalle diario por factura."""
    anchos = {
        "A": 6,
        "B": 12,
        "C": 18,
        "D": 28,
        "E": 8,
        "F": 14,
    }
    for letra, ancho in anchos.items():
        hoja.column_dimensions[letra].width = ancho


def _escribir_totales_reporte_diario(hoja, fila_inicio: int, reporte: Dict[str, Any]) -> None:
    """Escribe totales generales y por método de pago al final del Excel diario."""
    totales_metodo = reporte.get("totales_por_metodo_pago", {})
    fila = fila_inicio
    resumen = [
        ("Total", reporte["total_ventas"]),
    ]
    for codigo in ("anotar", "nequi", "daviplata", "efectivo"):
        etiqueta = ETIQUETAS_METODO_PAGO.get(codigo, codigo)
        resumen.append((f"Total {etiqueta}", int(totales_metodo.get(codigo, 0))))
    resumen.append(("Número de facturas", reporte["numero_facturas"]))

    for etiqueta, valor in resumen:
        hoja.cell(row=fila, column=5, value=etiqueta)
        hoja.cell(row=fila, column=5).font = Font(
            bold=True, color=MARCA_COLORES["naranja_oscuro"].lstrip("#")
        )
        hoja.cell(row=fila, column=6, value=valor)
        hoja.cell(row=fila, column=6).font = Font(
            bold=True, color=MARCA_COLORES["naranja_oscuro"].lstrip("#")
        )
        fila += 1


def exportar_reporte_diario_excel(
    reporte: Dict[str, Any],
    ruta_destino: Optional[Path] = None,
) -> Path:
    """
    Genera un Excel del reporte diario con detalle por factura y totales por método.
    Retorna la ruta del archivo guardado.
    """
    fecha = reporte["fecha"]
    if ruta_destino is None:
        RUTA_EXPORTACION.mkdir(parents=True, exist_ok=True)
        ruta_destino = RUTA_EXPORTACION / f"reporte_diario_{fecha}.xlsx"
    _asegurar_directorio_destino(ruta_destino)

    libro = Workbook()
    hoja = libro.active
    hoja.title = _NOMBRE_HOJA

    fila_datos = _escribir_encabezado_hoja(
        hoja,
        "Reporte de ventas diario",
        f"Fecha: {fecha}",
    )

    encabezados = list(ENCABEZADOS_DETALLE_DIARIO)
    for col, texto in enumerate(encabezados, start=1):
        hoja.cell(row=fila_datos, column=col, value=texto)
    _aplicar_estilo_encabezado_tabla(hoja, fila_datos, len(encabezados))

    fila_actual = fila_datos + 1
    detalle = reporte.get("detalle_ventas", [])
    if detalle:
        for entrada in detalle_con_separadores_factura(detalle):
            if entrada is None:
                hoja.row_dimensions[fila_actual].height = 8
                fila_actual += 1
                continue
            hoja.cell(
                row=fila_actual,
                column=1,
                value=numero_factura_corto(entrada["factura_numero"]),
            )
            hoja.cell(
                row=fila_actual,
                column=2,
                value=etiqueta_metodo_pago(entrada["metodo_pago"]),
            )
            hoja.cell(
                row=fila_actual,
                column=3,
                value=texto_comprador(entrada["comprador_nombre"]),
            )
            hoja.cell(row=fila_actual, column=4, value=entrada["nombre_producto"])
            hoja.cell(row=fila_actual, column=5, value=entrada["cantidad"])
            hoja.cell(row=fila_actual, column=6, value=entrada["subtotal"])
            fila_actual += 1
    else:
        hoja.cell(row=fila_actual, column=4, value="Sin ventas registradas")
        fila_actual += 1

    fila_total = fila_actual + 1
    hoja.cell(row=fila_total, column=5, value="TOTAL SUBTOTALES")
    hoja.cell(row=fila_total, column=5).font = Font(
        bold=True, color=MARCA_COLORES["naranja_oscuro"].lstrip("#")
    )
    col_subtotal = get_column_letter(6)
    primera_fila = fila_datos + 1
    ultima_fila = fila_actual - 1
    if detalle:
        hoja.cell(
            row=fila_total,
            column=6,
            value=f"=SUM({col_subtotal}{primera_fila}:{col_subtotal}{ultima_fila})",
        )
    else:
        hoja.cell(row=fila_total, column=6, value=0)
    hoja.cell(row=fila_total, column=6).font = Font(
        bold=True, color=MARCA_COLORES["naranja_oscuro"].lstrip("#")
    )

    _escribir_totales_reporte_diario(hoja, fila_total + 2, reporte)

    _ajustar_columnas_detalle_diario(hoja)
    _ajustar_ancho_columnas(hoja)
    libro.save(str(ruta_destino))
    return Path(ruta_destino)


def exportar_reporte_mensual_excel(
    reporte: Dict[str, Any],
    ruta_destino: Optional[Path] = None,
) -> Path:
    """
    Genera un Excel del reporte mensual en hoja Ventas con fila SUM() de ventas.
    Retorna la ruta del archivo guardado.
    """
    anio = reporte["anio"]
    mes = reporte["mes"]
    if ruta_destino is None:
        RUTA_EXPORTACION.mkdir(parents=True, exist_ok=True)
        ruta_destino = RUTA_EXPORTACION / f"reporte_mensual_{anio}-{mes:02d}.xlsx"
    _asegurar_directorio_destino(ruta_destino)

    libro = Workbook()
    hoja = libro.active
    hoja.title = _NOMBRE_HOJA

    fila_datos = _escribir_encabezado_hoja(
        hoja,
        "Reporte de ventas mensual",
        f"Período: {mes:02d}/{anio}",
    )

    encabezados = ["Fecha", "Total ventas (COP)", "Nº facturas"]
    for col, texto in enumerate(encabezados, start=1):
        hoja.cell(row=fila_datos, column=col, value=texto)
    _aplicar_estilo_encabezado_tabla(hoja, fila_datos, len(encabezados))

    fila_actual = fila_datos + 1
    cierres = reporte.get("cierres_diarios", [])
    if cierres:
        for cierre in cierres:
            hoja.cell(row=fila_actual, column=1, value=cierre["fecha"])
            hoja.cell(row=fila_actual, column=2, value=cierre["total_ventas"])
            hoja.cell(row=fila_actual, column=3, value=cierre["numero_facturas"])
            fila_actual += 1
    else:
        hoja.cell(row=fila_actual, column=1, value="Sin cierres registrados")
        fila_actual += 1

    fila_total = fila_actual
    hoja.cell(row=fila_total, column=1, value="TOTAL")
    hoja.cell(row=fila_total, column=1).font = Font(
        bold=True, color=MARCA_COLORES["naranja_oscuro"].lstrip("#")
    )
    col_ventas = get_column_letter(2)
    col_facturas = get_column_letter(3)
    primera_fila = fila_datos + 1
    ultima_fila = fila_actual - 1
    if cierres:
        hoja.cell(
            row=fila_total,
            column=2,
            value=f"=SUM({col_ventas}{primera_fila}:{col_ventas}{ultima_fila})",
        )
        hoja.cell(
            row=fila_total,
            column=3,
            value=f"=SUM({col_facturas}{primera_fila}:{col_facturas}{ultima_fila})",
        )
    else:
        hoja.cell(row=fila_total, column=2, value=0)
        hoja.cell(row=fila_total, column=3, value=0)
    hoja.cell(row=fila_total, column=2).font = Font(
        bold=True, color=MARCA_COLORES["naranja_oscuro"].lstrip("#")
    )
    hoja.cell(row=fila_total, column=3).font = Font(
        bold=True, color=MARCA_COLORES["naranja_oscuro"].lstrip("#")
    )

    _ajustar_ancho_columnas(hoja)
    libro.save(str(ruta_destino))
    return Path(ruta_destino)
